import warnings
import string
from itertools import product

import numpy as np
import pandas as pd
from statsmodels.stats.proportion import proportions_ztest

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


letters_list = list(string.ascii_uppercase)

red_fill = PatternFill(start_color="C80000", end_color="C80000", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
blue_fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")


class ExcelWriter:
    def __init__(self, xlsx_file: str):
        self.xlsx_file = xlsx_file
        self.workbook = load_workbook(xlsx_file)
        self.index_totals = 1

    def copy_styles(self, cell_source, cell_target):
        if cell_source.has_style:
            cell_target.font = Font(
                name=cell_source.font.name,
                size=cell_source.font.size,
                bold=cell_source.font.bold,
                italic=cell_source.font.italic,
                vertAlign=cell_source.font.vertAlign,
                underline=cell_source.font.underline,
                strike=cell_source.font.strike,
                color=cell_source.font.color,
            )

            cell_target.border = Border(
                left=Side(
                    border_style=cell_source.border.left.style,
                    color=cell_source.border.left.color,
                ),
                right=Side(
                    border_style=cell_source.border.right.style,
                    color=cell_source.border.right.color,
                ),
                top=Side(
                    border_style=cell_source.border.top.style,
                    color=cell_source.border.top.color,
                ),
                bottom=Side(
                    border_style=cell_source.border.bottom.style,
                    color=cell_source.border.bottom.color,
                ),
            )

            cell_target.fill = PatternFill(
                fill_type=cell_source.fill.fill_type,
                start_color=cell_source.fill.start_color,
                end_color=cell_source.fill.end_color,
            )

            cell_target.number_format = cell_source.number_format
            cell_target.protection = Protection(
                locked=cell_source.protection.locked,
                hidden=cell_source.protection.hidden,
            )
            cell_target.alignment = Alignment(
                horizontal=cell_source.alignment.horizontal,
                vertical=cell_source.alignment.vertical,
                text_rotation=cell_source.alignment.text_rotation,
                wrap_text=cell_source.alignment.wrap_text,
                shrink_to_fit=cell_source.alignment.shrink_to_fit,
                indent=cell_source.alignment.indent,
            )

    def apply_red_color_to_letter(self, cell):
        value = cell.value
        if isinstance(value, str) and any(char.isalpha() for char in value):
            index = 0
            while index < len(value) and value[index].isdigit():
                index += 1

            num = value[:index]
            letter = value[index:].strip()

            red = InlineFont(color="00FF0000")
            rich_text_cell = CellRichText()
            rich_text_cell.append(f"{num} ")
            rich_text_cell.append(TextBlock(red, letter))
            cell.value = rich_text_cell

    def write_penalty_sheet(self, result_df: pd.DataFrame, worksheet: Worksheet):
        unique_questions = result_df["question"].unique().tolist()
        dfs = {
            question: result_df[result_df["question"] == question].copy()
            for question in unique_questions
        }

        worksheet.sheet_view.showGridLines = False

        start_row = 0
        for question, df in dfs.items():
            headers = ["Grouped Variable"] + df.columns[2:].to_list()
            for col_num, header in enumerate(headers, 2):
                cell = worksheet.cell(row=start_row + 1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.border = Border(bottom=Side(border_style="thin"))
                cell.alignment = Alignment(horizontal="center")

            for row_num, row_data in enumerate(
                dataframe_to_rows(df, index=False, header=False), start=start_row + 2
            ):
                for col_num, cell_value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    if col_num > 2:
                        cell.number_format = "0.00"
                    if col_num > 1:
                        cell.border = Border(
                            top=Side(border_style="thin"),
                            bottom=Side(border_style="thin"),
                        )

            cell = worksheet.cell(row=start_row + 2, column=1)
            cell.value = question
            cell.alignment = Alignment(vertical="top", wrapText=True)
            worksheet.merge_cells(
                start_row=start_row + 2,
                start_column=1,
                end_row=start_row + 1 + len(df),
                end_column=1,
            )

            for col in range(1, len(df.columns) + 1):
                worksheet.cell(row=start_row + 1, column=col).border = Border(
                    top=Side(border_style="thick")
                )
                worksheet.cell(row=start_row + 1 + len(df), column=col).border = Border(
                    bottom=Side(border_style="thick")
                )

            worksheet.column_dimensions["A"].width = 400 / 8.43
            worksheet.column_dimensions["B"].width = 250 / 8.43

            start_row += len(df) + 3

    def write_statistical_significance_sheet(
        self,
        existing_worksheet,
        new_worksheet,
        first_all_nan_index,
        combined_differences_df,
        question_groups,
        category_indexes,
    ):
        for row in existing_worksheet.iter_rows():
            for cell in row:
                new_cell = new_worksheet.cell(
                    row=cell.row, column=cell.column, value=cell.value
                )
                self.copy_styles(cell, new_cell)

        for col in existing_worksheet.columns:
            column_letter = col[first_all_nan_index + 1].column_letter
            new_worksheet.column_dimensions[
                column_letter
            ].width = existing_worksheet.column_dimensions[column_letter].width

        start_row = 1
        start_column = 1
        for r_idx, row in enumerate(
            dataframe_to_rows(combined_differences_df, index=False, header=True),
            start=start_row,
        ):
            for c_idx, value in enumerate(row, start=start_column):
                if not (isinstance(value, str) and "Unnamed" in value):
                    new_cell = new_worksheet.cell(row=r_idx, column=c_idx, value=value)

        for merged_cell in existing_worksheet.merged_cells.ranges:
            new_worksheet.merge_cells(str(merged_cell))

        for merged_range in list(new_worksheet.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(merged_range.coord)

            if min_col == 2 and max_col == 2:
                new_worksheet.unmerge_cells(
                    start_row=min_row,
                    start_column=min_col,
                    end_row=max_row,
                    end_column=max_col,
                )

        for row in new_worksheet.iter_rows():
            for cell in row[:3]:
                cell.font = Font(bold=True)

        self.delete_col_with_merged_ranges(new_worksheet, 2)

        for question_group in question_groups:
            for row in question_group:
                for category_group in category_indexes:
                    col_start, col_end = category_group
                    for col in range(col_start, col_end + 1):
                        cell = new_worksheet.cell(row=row + 2, column=col)
                        self.apply_red_color_to_letter(cell)

        new_worksheet.cell(row=1, column=1).value = ""

        new_worksheet.column_dimensions["A"].width = 400 / 8.43
        new_worksheet.column_dimensions["B"].width = 150 / 8.43

        fixed_column_width = 80 / 8.43
        for col in range(3, new_worksheet.max_column + 1):
            column_letter = get_column_letter(col)
            new_worksheet.column_dimensions[column_letter].width = fixed_column_width

        fixed_row_height = 20 / 1.33
        for row in range(1, new_worksheet.max_row + 1):
            new_worksheet.row_dimensions[row].height = fixed_row_height

        for row in new_worksheet.iter_rows(min_col=2):
            for cell in row:
                if cell.alignment.wrap_text:
                    alignment = cell.alignment
                    new_alignment = Alignment(
                        horizontal=alignment.horizontal,
                        vertical=alignment.vertical,
                        text_rotation=alignment.text_rotation,
                        wrap_text=False,
                        shrink_to_fit=alignment.shrink_to_fit,
                        indent=alignment.indent,
                        justifyLastLine=alignment.justifyLastLine,
                        readingOrder=alignment.readingOrder,
                    )
                    cell.alignment = new_alignment

    def delete_row_with_merged_ranges(self, sheet, idx):
        sheet.delete_rows(idx)
        for mcr in sheet.merged_cells:
            if idx < mcr.min_row:
                mcr.shift(row_shift=-1)
            elif idx < mcr.max_row:
                mcr.shrink(bottom=1)

    def delete_col_with_merged_ranges(self, sheet, idx):
        sheet.delete_cols(idx)
        for mcr in sheet.merged_cells:
            if idx < mcr.min_col:
                mcr.shift(col_shift=-1)
            elif idx < mcr.max_col:
                mcr.shrink(right=1)

    def process_netos(self, wstemp):
        maxcol = wstemp.max_column
        dictionary_netos = {}

        # What is the purpose of this loop?
        for rowi in range(1, wstemp.max_row + 1):
            valb = wstemp["B" + str(rowi)].value
            if (
                valb
                and valb.startswith("NETO")
                and valb != "NETO TOP TWO BOX"
                and valb != "NETO BOTTOM TWO BOX"
            ):
                if valb not in dictionary_netos or dictionary_netos[valb] == 0:
                    dictionary_netos[valb] = 1
                    for rowf in range(rowi + 1, wstemp.max_row + 1):
                        if valb == wstemp["B" + str(rowf)].value:
                            for i in range(2, maxcol + 1):
                                wstemp[get_column_letter(i) + str(rowi)] = wstemp[
                                    get_column_letter(i) + str(rowf)
                                ].value
                            for merged_range in list(wstemp.merged_cells.ranges):
                                if (
                                    merged_range.min_row >= rowf - 7
                                    and merged_range.min_row <= rowf + 5
                                ):
                                    wstemp.merged_cells.ranges.remove(merged_range)
                            break
                else:
                    dictionary_netos[valb] = dictionary_netos[valb] - 1

        # What is the purpose of this loop?
        for rowi in range(1, wstemp.max_row + 1):
            valb = wstemp["B" + str(rowi)].value
            if (
                valb
                and valb.startswith("NETO")
                and valb != "NETO TOP TWO BOX"
                and valb != "NETO BOTTOM TWO BOX"
            ):
                for rowf in range(rowi + 1, wstemp.max_row + 1):
                    if valb == wstemp["B" + str(rowf)].value:
                        for _ in range(11):
                            self.delete_row_with_merged_ranges(wstemp, rowf - 7)
                        break

        # What is the purpose of this loop?
        for rowi in range(wstemp.max_row + 1, 1, -1):
            if not wstemp["D" + str(rowi)].value and not wstemp["C" + str(rowi)].value:
                self.delete_row_with_merged_ranges(wstemp, rowi)

    def preformat_sheets(self):
        # Preformat the existing sheets
        for sheet in self.workbook:
            if not sheet.title.lower().startswith("penal"):
                self.process_netos(self.workbook[sheet.title])

        self.workbook.save(self.xlsx_file)

    def format_columns(self, ws_totals: Worksheet):
        separators = []

        for col in range(1, ws_totals.max_column + 1):
            column_letter = get_column_letter(col)

            actual_cell_value = ws_totals.cell(row=1, column=col).value
            next_cell_value = ws_totals.cell(row=1, column=col + 1).value

            if actual_cell_value is None and next_cell_value is None:
                ws_totals.column_dimensions[column_letter].width = 3

            if actual_cell_value is None and next_cell_value is not None:
                ws_totals.column_dimensions[column_letter].width = 4
                separators.append(col)

            if actual_cell_value is not None:
                ws_totals.column_dimensions[column_letter].width = 14

        for col in separators:
            for i in range(1, ws_totals.max_row + 1):
                ws_totals.cell(row=i, column=col).fill = blue_fill

    def replicate_with_formatting(self, ws_totals: Worksheet, new_worksheet: Worksheet):
        max_col = new_worksheet.max_column
        index_row = 2
        index_ini = 0
        for row_actual in range(1, new_worksheet.max_row + 1):
            val_a = new_worksheet["A" + str(row_actual)].value
            val_b = new_worksheet["B" + str(row_actual)].value
            if val_a:
                ws_totals.cell(row=index_row, column=self.index_totals).value = val_a
            elif val_b == "Total":
                if index_ini == 0:
                    index_ini = row_actual

                for i in range(2, max_col):
                    row_actual_cell = new_worksheet.cell(row=row_actual, column=i + 1)
                    index_ini_cell = new_worksheet.cell(row=index_ini, column=i + 1)
                    previous_cell = ws_totals.cell(
                        row=index_row, column=self.index_totals + i - 1
                    )

                    previous_cell.value = row_actual_cell.value

                    if row_actual_cell.value is np.nan:
                        previous_cell.fill = red_fill
                    elif row_actual_cell.value != index_ini_cell.value:
                        previous_cell.fill = yellow_fill

                index_row += 1

        self.index_totals += max_col


class DataProcessor:
    @staticmethod
    def extract_digits(cell):
        cell = str(cell)
        match = pd.Series(cell).str.extract("(\d+)")[0][0]
        return int(match) if pd.notna(match) else None

    @staticmethod
    def calculate_percentages(
        inner_df: pd.DataFrame, total_df: pd.DataFrame, total_index: int
    ) -> pd.DataFrame:
        percentage_inner_df = inner_df.copy()
        for column in inner_df.columns:
            percentage_inner_df[column] = (
                inner_df[column] / total_df.loc[total_index, column]
            ) * 100
        return percentage_inner_df

    @staticmethod
    def group_consecutive_indexes(index_list: list):
        if not index_list:
            return []

        sorted_indexes = sorted(set(index_list))
        groups = []
        group = [sorted_indexes[0]]

        for i in range(1, len(sorted_indexes)):
            if sorted_indexes[i] == sorted_indexes[i - 1] + 1:
                group.append(sorted_indexes[i])
            else:
                groups.append(group)
                group = [sorted_indexes[i]]

        groups.append(group)
        return groups

    @staticmethod
    def column_to_numeric(column: str, data: pd.DataFrame) -> pd.DataFrame:
        data[column] = (
            pd.to_numeric(data[column], errors="coerce").fillna(data[column]).fillna("")
        )
        return data

    @staticmethod
    def transform_headers(data: pd.DataFrame):
        c = 0
        new_columns = []
        for column in data.loc[0, :]:
            if pd.isnull(column):
                new_columns.append(f"Unnamed: {c}")
                c += 1
            else:
                new_columns.append(column)

        data.columns = new_columns
        data = data.drop(0).reset_index(drop=True)
        return data

    @staticmethod
    def composite_columns(n: int, start: str = "AA"):
        letters = string.ascii_uppercase
        start_index = letters.index(start[0]) * 26 + letters.index(start[1]) + 1
        columns = ["".join(pair) for pair in product(letters, repeat=2)]
        return columns[start_index - 1 : start_index - 1 + n]

    @staticmethod
    def calculate_differences(
        x1: int, x2: int, n1: int, n2: int, sigma: float = 0.05
    ) -> bool:
        if n1 < 30 or n2 < 30 or x1 == 0 or x2 == 0 or n1 == 0 or n2 == 0:
            return False

        counts = np.array([x1, x2])
        nobs = np.array([n1, n2])

        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=RuntimeWarning)
            _, p_value = proportions_ztest(counts, nobs)

        return p_value < sigma

    @staticmethod
    def statistical_significance(
        inner_df: pd.DataFrame,
        data: pd.DataFrame,
        total_index: int,
        letters_inner_dict: dict[str, str],
    ) -> pd.DataFrame:
        columns = inner_df.columns
        n_cols = len(columns)

        inner_differences_df = pd.DataFrame(
            "", index=inner_df.index, columns=inner_df.columns
        )

        for i in range(n_cols):
            for j in range(i + 1, n_cols):
                col1 = columns[i]
                col2 = columns[j]

                for index in inner_df.index:
                    x1 = inner_df.at[index, col1]
                    x2 = inner_df.at[index, col2]
                    n1 = data.loc[total_index, col1]
                    n2 = data.loc[total_index, col2]

                    if DataProcessor.calculate_differences(x1, x2, n1, n2):
                        if (x1 / n1) > (x2 / n2):
                            if inner_differences_df.at[index, col1]:
                                inner_differences_df.at[index, col1] += (
                                    f",{letters_inner_dict[col2]}"
                                )
                            else:
                                inner_differences_df.at[index, col1] = (
                                    letters_inner_dict[col2]
                                )
                        else:
                            if inner_differences_df.at[index, col2]:
                                inner_differences_df.at[index, col2] += (
                                    f",{letters_inner_dict[col1]}"
                                )
                            else:
                                inner_differences_df.at[index, col2] = (
                                    letters_inner_dict[col1]
                                )

        return inner_differences_df

    @staticmethod
    def combine_values(num: int | float, string: str, decimals: int = 2):
        if pd.isna(num) and pd.isna(string):
            return np.nan
        elif pd.isna(num):
            return string.strip()
        elif pd.isna(string):
            if isinstance(num, float):
                return f"{num:.{decimals}f}".strip()
            else:
                return str(num).strip()
        else:
            if isinstance(num, float):
                num = f"{num:.{decimals}f}"
            return f"{num} {string}".strip()

    @staticmethod
    def combine_dataframes(df1, df2, decimals=2):
        combined_data = {
            col: [
                DataProcessor.combine_values(num, string, decimals)
                for num, string in zip(df1[col], df2[col])
            ]
            for col in df1.columns
        }

        combined_df = pd.concat(
            [pd.Series(values, name=col) for col, values in combined_data.items()],
            axis=1,
        )

        return combined_df

    @staticmethod
    def extract_statistical_significance_metadata(data: pd.DataFrame):
        float_types = data["Unnamed: 2"].apply(lambda x: isinstance(x, float))
        float_types = float_types[float_types]

        index_list = list(float_types.index)

        question_groups = DataProcessor.group_consecutive_indexes(index_list)

        partial_df = data[
            data.index.isin([question_groups[0][0] - 1] + question_groups[0])
        ]
        initial_category_group = partial_df.loc[partial_df.index[0], "TOTAL"]  # (A)
        category_groups_columns = (
            partial_df.loc[partial_df.index[0], :].to_frame().reset_index()
        )

        initial_category_indexes = DataProcessor.group_consecutive_indexes(
            list(
                category_groups_columns[
                    category_groups_columns[1] == initial_category_group
                ][1:].index
            )
        )

        category_indexes = [
            (
                initial_category_indexes[i][-1],
                initial_category_indexes[i + 1][0] - 1,
            )
            for i in range(len(initial_category_indexes) - 1)
        ] + [(initial_category_indexes[-1][0], len(category_groups_columns) - 1)]

        for cat in initial_category_indexes:
            if len(cat) > 1:
                for cat1 in cat:
                    if cat1 != cat[-1]:
                        category_indexes += [(cat1, cat1)]

        return question_groups, category_indexes, category_groups_columns

    @staticmethod
    def process_statistical_significance(
        data: pd.DataFrame, question_groups, category_indexes, category_groups_columns
    ):
        data = DataProcessor.column_to_numeric("Unnamed: 2", data)

        data_statistical_significance = data.copy()

        total_statistical_significance_df = pd.DataFrame(
            index=range(len(data_statistical_significance)),
            columns=data_statistical_significance.columns,
        )

        for question_group in question_groups:
            df_total_search = data_statistical_significance.loc[
                question_group[-1] : question_group[-1] + 6, :
            ]
            total_index = df_total_search[
                df_total_search["Unnamed: 2"].str.contains("Total", na=False)
            ].index[0]
            data_statistical_significance.loc[total_index, "TOTAL"] = int(
                data_statistical_significance.loc[total_index, "TOTAL"]
            )

            data_statistical_significance.update(
                DataProcessor.calculate_percentages(
                    data_statistical_significance[["TOTAL"]]
                    .loc[question_group, :]
                    .astype(int),
                    data_statistical_significance,
                    total_index,
                )
            )

            for category_group in category_indexes:
                columns_category_groups = category_groups_columns.loc[
                    category_group[0] : category_group[1]
                ]["index"].to_list()

                inner_df = (
                    data.loc[question_group, columns_category_groups]
                    .map(DataProcessor.extract_digits)
                    .replace({None: np.nan})
                    .dropna(axis=1, how="all")
                )

                data_statistical_significance.update(inner_df)

                data_statistical_significance.loc[total_index, inner_df.columns] = (
                    data_statistical_significance.loc[total_index, inner_df.columns]
                    .infer_objects(copy=False)
                    .fillna(0)
                    .astype(int)
                )

                data_statistical_significance.update(
                    DataProcessor.calculate_percentages(
                        inner_df, data_statistical_significance, total_index
                    )
                )

                if len(inner_df.columns) > len(letters_list):
                    letters_inner_dict = {
                        column: letter
                        for column, letter in zip(
                            inner_df.columns,
                            DataProcessor.composite_columns(len(inner_df.columns)),
                        )
                    }
                else:
                    letters_inner_dict = {
                        column: letter
                        for column, letter in zip(
                            inner_df.columns, letters_list[: len(inner_df.columns)]
                        )
                    }

                inner_differences_df = DataProcessor.statistical_significance(
                    inner_df,
                    data_statistical_significance,
                    total_index,
                    letters_inner_dict,
                )

                total_statistical_significance_df.update(inner_differences_df)

        combined_differences_df = DataProcessor.combine_dataframes(
            data_statistical_significance, total_statistical_significance_df, 0
        )

        combined_differences_df["Unnamed: 2"] = np.where(
            combined_differences_df["Unnamed: 2"] == "1",
            combined_differences_df["Unnamed: 1"],
            combined_differences_df["Unnamed: 2"],
        )

        combined_differences_df["Unnamed: 2"] = combined_differences_df[
            "Unnamed: 2"
        ].replace("", np.nan)

        return combined_differences_df

    @staticmethod
    def extract_penalty_metadata(data: pd.DataFrame):
        first_row_with_data = data[~data.iloc[:, 3].isna()].index[0]
        data = data.iloc[first_row_with_data:]
        data.columns = [
            f"Unnamed: {i}" if pd.isna(col) else col
            for i, col in enumerate(data.iloc[0])
        ]
        data = data[1:]
        data = data.reset_index(drop=True)

        # Create a dataframe with the column names as the first row
        column_names_df = pd.DataFrame(
            [
                [
                    column if not column.startswith("Unnamed") else np.nan
                    for column in data.columns
                ]
            ],
            columns=list(range(len(data.columns))),
        )

        data.columns = range(len(data.columns))
        # Concatenate the new row with the original dataframe
        data = pd.concat([column_names_df, data]).reset_index(drop=True)

        # Rename the columns to integers
        data.columns = range(len(data.columns))

        questions = (
            data[data[0].str.startswith("P", na=False)][0].dropna().unique().tolist()
        )

        question_idexes = np.array(
            [data[data[0] == question].first_valid_index() for question in questions]
        )

        tables_first_indexes = (question_idexes - 2).tolist()

        tables_last_indexes = [
            data.iloc[start_idx + 2 :][
                (data[1].iloc[start_idx + 2 :].isna())
                & (data[2].iloc[start_idx + 2 :].isna())
            ].index[0]
            for start_idx in tables_first_indexes
        ]

        samples = data.loc[1, 3:].values.tolist()
        data.columns = ["question", "grouped_variable", "answer_option"] + samples

        tables_range_indexes = list(zip(tables_first_indexes, tables_last_indexes))

        return questions, tables_range_indexes, samples

    @staticmethod
    def process_penalty_samples(grouped_variables, samples, sub_df, base_inner_df):
        for sample in samples:
            total = np.nan
            for grouped_variable in grouped_variables:
                grouped_variable_index_df = sub_df[
                    sub_df["grouped_variable"] == grouped_variable
                ]
                if grouped_variable_index_df.empty:
                    continue
                grouped_variable_index = grouped_variable_index_df.index[0]
                grouped_variable_df = sub_df.loc[
                    grouped_variable_index : grouped_variable_index + 4
                ]

                if grouped_variable_df[sample].sum() == 0:
                    continue

                total = sub_df[:1][sample].values[0]

                base_inner_df.loc[grouped_variable, sample] = (
                    grouped_variable_df[sample].sum() / total
                )  # Percentage

                base_inner_df.loc[f"MEAN {grouped_variable} VS. IC", sample] = (
                    grouped_variable_df[sample] * np.array(list(range(0, 101, 25)))
                ).sum() / grouped_variable_df[sample].sum()  # Mean

            for grouped_variable in grouped_variables:
                if "just" not in grouped_variable.lower() and total is not np.nan:
                    percentage = base_inner_df.loc[grouped_variable, sample]
                    mean = base_inner_df.loc[f"MEAN {grouped_variable} VS. IC", sample]
                    jr_mean = base_inner_df.loc[
                        f"MEAN {grouped_variables[1]} VS. IC", sample
                    ]

                    base_inner_df.loc[f"PENALTY {grouped_variable}", sample] = (
                        mean - jr_mean
                    ) * percentage  # Penalty

            base_inner_df.loc["TOTAL", sample] = total  # Total

        return base_inner_df

    @staticmethod
    def process_penalty_data(data: pd.DataFrame) -> pd.DataFrame:
        questions, tables_range_indexes, samples = (
            DataProcessor.extract_penalty_metadata(data)
        )

        results_dfs = []

        for question, (start, end) in zip(questions, tables_range_indexes):
            question_df = data.loc[start:end, :]

            # Finding the index of the first occurrence
            first_occurrence_index = question_df[
                question_df["grouped_variable"].str.contains("Total", na=False)
            ].index[0]

            grouped_variables = (
                question_df.loc[: first_occurrence_index - 1]
                .dropna(subset="grouped_variable")["grouped_variable"]
                .to_list()
            )

            results_calculations = (
                grouped_variables
                + [
                    f"MEAN {grouped_variable} VS. IC"
                    for grouped_variable in grouped_variables
                ]
                + [
                    f"PENALTY {grouped_variable}"
                    for grouped_variable in grouped_variables
                    if "just" not in grouped_variable.lower()
                ]
                + ["TOTAL"]
            )

            base_inner_df = pd.DataFrame(index=results_calculations)

            sub_df = question_df.loc[first_occurrence_index:]

            base_inner_df = DataProcessor.process_penalty_samples(
                grouped_variables, samples, sub_df, base_inner_df
            )

            base_inner_df = base_inner_df.reset_index(names="grouped_variable")
            base_inner_df.insert(0, "question", question)

            results_dfs.append(base_inner_df)

        result_df = pd.concat(results_dfs).reset_index(drop=True)

        return result_df


def calculate_statistical_significance(xlsx_file: str):
    # Load the existing Excel file
    excel_writer = ExcelWriter(xlsx_file)
    excel_writer.preformat_sheets()

    # Create a new Workbook
    new_workbook = Workbook()

    # Remove the default sheet created with the new workbook
    default_sheet = new_workbook.active
    new_workbook.remove(default_sheet)

    sheets_dfs = pd.read_excel(xlsx_file, sheet_name=None)

    totals_worksheet = new_workbook.create_sheet(title="TOTALES")

    # Iterate over all sheets
    for sheet_name, data in sheets_dfs.items():
        if data.empty:
            continue

        new_worksheet = new_workbook.create_sheet(title=sheet_name)

        # Process and write the penalty data
        if sheet_name.lower().startswith("penal"):
            result_df = DataProcessor.process_penalty_data(data)
            excel_writer.write_penalty_sheet(result_df, new_worksheet)

        else:
            existing_worksheet = excel_writer.workbook[sheet_name]

            if "TOTAL" not in data.columns:
                data = DataProcessor.transform_headers(data)
                excel_writer.delete_row_with_merged_ranges(existing_worksheet, 0)

            data = DataProcessor.column_to_numeric("Unnamed: 2", data)

            question_groups, category_indexes, category_groups_columns = (
                DataProcessor.extract_statistical_significance_metadata(data)
            )

            combined_statistical_significance_df = (
                DataProcessor.process_statistical_significance(
                    data, question_groups, category_indexes, category_groups_columns
                )
            )

            nan_df = combined_statistical_significance_df[
                combined_statistical_significance_df.isna().all(axis=1)
            ]

            if not nan_df.empty:
                first_all_nan_index = combined_statistical_significance_df[
                    combined_statistical_significance_df.isna().all(axis=1)
                ].index[0]
            else:
                first_all_nan_index = 2

            excel_writer.write_statistical_significance_sheet(
                existing_worksheet,
                new_worksheet,
                first_all_nan_index,
                combined_statistical_significance_df,
                question_groups,
                category_indexes,
            )

            totals_worksheet.cell(
                row=1, column=excel_writer.index_totals
            ).value = sheet_name

            excel_writer.replicate_with_formatting(existing_worksheet, new_worksheet)

    excel_writer.format_columns(totals_worksheet)

    output_xlsx_file = xlsx_file.replace(".xlsx", "_processed.xlsx")
    excel_writer.workbook.save(output_xlsx_file)

    return output_xlsx_file
